"""
LEAP Publication Scraper  v4.0
==============================
Searches multiple academic databases for publications affiliated with
LEAP (Learning the Earth with AI and Physics), NSF Award #2019625.

Data Sources:
  1. OpenAlex        – broad academic index, author + keyword search
  2. CrossRef        – funder search by NSF award number (#2019625)
  3. NASA ADS        – earth/atmospheric science coverage (requires free API key)
  4. ORCID           – author publication lists (for scientists with ORCID IDs)
  5. arXiv           – preprints (AI + physics + climate)
  6. Semantic Scholar – supplementary AI-focused index

Full-Text PDF Verification (v4):
  After collecting papers, the script downloads the full PDF for every paper
  marked "Likely" or "No" and searches the ENTIRE document for LEAP signals.
  No HTML scraping — PDFs are consistent across all publishers.

  PDF sources tried in order:
    1. Unpaywall API        → finds OA PDF for any DOI (free, no key needed)
    2. Publisher direct     → arXiv, EarthArXiv, ESSOAr, Copernicus, PLOS
    3. DOI link itself      → sometimes resolves directly to a PDF

  Signals searched (case-insensitive, whole document):
    "learning the earth with ai and physics", "leap science technology center",
    "leap center", "leap stc", "2019625", "ags-2019625", "nsf stc", etc.

  Papers where any signal is found → upgraded to "Yes".
  Notes column records exactly which signal matched and which PDF URL.

Configuration:
  - Scientists loaded from: leap_scientists.txt
  - Start date filter:      START_DATE (default May 1, 2025)
  - NSF Award:              #2019625
  - Output:                 leap_publications.xlsx saved to gs://leap-persistent/publications/

NASA ADS Setup (free):
  1. Register at https://ui.adsabs.harvard.edu/user/account/register
  2. Get your API token at https://ui.adsabs.harvard.edu/user/settings/token
  3. Run: export ADS_API_TOKEN=your_token_here

Usage:
  pip install requests openpyxl pandas pdfplumber beautifulsoup4
  python leap_publication_scraper.py
"""

# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "requests",
#   "pandas",
#   "openpyxl",
#   "pdfplumber",
#   "beautifulsoup4",
# ]
# /// Imports and configuration

import os
import re
import io
import subprocess
import time
import requests
import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from leap_batch_jobs.monitoring import ProgressLogger, ResourceMonitor, notify_slack

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("Note: pdfplumber not installed – PDF extraction disabled. Run: pip install pdfplumber")

# ─── USER CONFIG ───────────────────────────────────────────────────────────────

SCIENTISTS_FILE = "leap_scientists.txt"
OUTPUT_FILE     = "leap_publications.xlsx"

# Scrape publications from this date onward
START_DATE = date(2025, 5, 1)

# NSF Award number (without the #)
NSF_AWARD = "2019625"

# NASA ADS API token – get yours free at https://ui.adsabs.harvard.edu
# Leave empty "" to skip NASA ADS searches
ADS_API_TOKEN = os.environ.get("ADS_API_TOKEN", "")

# Contact email for API polite pools
CONTACT_EMAIL = "leap-publications@columbia.edu"

# ─── KEYWORD / QUERY LISTS ─────────────────────────────────────────────────────

LEAP_KEYWORDS = [
    "Learning the Earth with AI and Physics",
    f"NSF {NSF_AWARD}",
    "LEAP NSF STC climate",
    "LEAP center Columbia climate AI",
]

ADS_QUERIES = [
    f"grant:{NSF_AWARD}",
    "LEAP NSF climate AI",
    "Learning the Earth AI Physics climate",
]

ARXIV_QUERIES = [
    'ti:LEAP AND (abs:climate OR abs:ocean OR abs:atmosphere) AND abs:"machine learning"',
    'abs:"Learning the Earth with AI and Physics"',
    f'abs:{NSF_AWARD}',
]

# ─── LOAD SCIENTISTS ───────────────────────────────────────────────────────────

def load_scientists(filepath):
    scientists = []
    if not os.path.exists(filepath):
        print(f"WARNING: Scientists file not found: {filepath}")
        return scientists
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "|" in line:
                name, orcid = line.split("|", 1)
                name  = name.strip()
                orcid = orcid.strip() or None
            else:
                name  = line.strip()
                orcid = None
            if name:
                scientists.append({"name": name, "orcid": orcid})
    print(f"  Loaded {len(scientists)} scientists from {filepath}")
    return scientists

# ─── SHARED UTILITIES ──────────────────────────────────────────────────────────

def safe_get(url, params=None, headers=None, retries=3, wait=2):
    _headers = {"User-Agent": f"LEAP-Scraper/2.0 (mailto:{CONTACT_EMAIL})"}
    if headers:
        _headers.update(headers)
    for attempt in range(retries):
        try:
            r = requests.get(url, params=params, headers=_headers, timeout=20)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            print(f"    Retry {attempt+1}/{retries} – {e}")
            time.sleep(wait * (attempt + 1))
    return None


def extract_last_name(full_name):
    if not full_name:
        return ""
    return full_name.strip().split()[-1]


def in_date_range(year_str, month_str=None):
    """Return True if year/month >= START_DATE."""
    try:
        year  = int(year_str)
        month = int(month_str) if month_str else 1
        return date(year, month, 1) >= START_DATE
    except Exception:
        return False


def detect_leap_ack(text):
    lower = (text or "").lower()
    signals = [
        "learning the earth with ai and physics",
        "nsf stc", "science and technology center",
        f"nsf {NSF_AWARD}", NSF_AWARD,
        "leap center", "leap stc",
    ]
    return "Yes" if any(s in lower for s in signals) else "No"

# ─── CITATION FORMATTER ────────────────────────────────────────────────────────

def fmt_author(name):
    """Convert 'First Last' or 'Last, First' -> 'Last F'."""
    name = (name or "").strip()
    if not name:
        return ""
    if "," in name:
        last, first = name.split(",", 1)
        initials = "".join(p[0] for p in first.strip().split() if p)
        return f"{last.strip()} {initials}"
    parts = name.split()
    if len(parts) == 1:
        return parts[0]
    last     = parts[-1]
    initials = "".join(p[0] for p in parts[:-1] if p)
    return f"{last} {initials}"


def format_citation(authors, title, journal, year, doi_link):
    """
    Produce NSF/LEAP citation style:
    Last FM, Last2 FM2 (YEAR) Title. Journal. DOI_URL
    """
    if authors:
        formatted = [fmt_author(a) for a in authors]
        if len(formatted) <= 6:
            author_str = ", ".join(formatted)
        else:
            author_str = ", ".join(formatted[:6]) + ", et al."
    else:
        author_str = ""

    year_str    = f"({year})" if year else ""
    title_str   = (title   or "").strip().rstrip(".")
    journal_str = (journal or "").strip().rstrip(".")
    doi_str     = doi_link or ""
    if doi_str and not doi_str.startswith("http"):
        doi_str = f"https://doi.org/{doi_str}"

    parts = [p for p in [author_str, year_str, title_str + ".", journal_str + "."] if p]
    base  = " ".join(parts)
    if doi_str:
        base += f" {doi_str}"
    return base


def make_row(last_name, title, link, journal, notes,
             leap_ack, year, month=None, authors=None):
    date_str = f"{year}-{int(month):02d}" if month else str(year or "")
    citation = format_citation(authors or [], title, journal, year, link)
    return {
        "Last Name":                  last_name or "",
        "Title":                      title     or "",
        "Link":                       link      or "",
        "Published In":               journal   or "",
        "Notes":                      notes     or "",
        "LEAP Funding Acknowledged?": leap_ack,
        "Date":                       date_str,
        "Citation (NSF Format)":      citation,
        "_work_id": (link or (title or "").lower().strip()[:80]),
        "_authors": authors or [],
    }

# ─── 1. OPENALEX ───────────────────────────────────────────────────────────────

def decode_abstract(inv):
    if not inv:
        return ""
    pairs = [(pos, w) for w, positions in inv.items() for pos in positions]
    return " ".join(w for _, w in sorted(pairs))


def parse_openalex(work, lastnames):
    title = work.get("title") or ""
    doi   = work.get("doi") or ""
    link  = doi if doi else work.get("id", "")
    year  = work.get("publication_year")
    month = None
    pub_date = work.get("publication_date") or ""
    if len(pub_date) >= 7:
        try:
            month = int(pub_date[5:7])
        except Exception:
            pass
    if not year or not in_date_range(str(year), str(month) if month else None):
        return None

    loc     = work.get("primary_location") or {}
    journal = (loc.get("source") or {}).get("display_name") or ""

    authorships = work.get("authorships") or []
    authors = [(a.get("author") or {}).get("display_name") or "" for a in authorships]
    first_last = extract_last_name(authors[0]) if authors else ""

    abstract = decode_abstract(work.get("abstract_inverted_index"))
    leap_ack = detect_leap_ack(f"{title} {abstract}")
    if leap_ack == "No":
        for a in authors:
            if extract_last_name(a).lower() in lastnames:
                leap_ack = "Likely"
                break

    notes = f"OpenAlex. Authors: {'; '.join(authors[:5])}{'...' if len(authors)>5 else ''}"
    return make_row(first_last, title, link, journal, notes, leap_ack, year, month, authors)


def openalex_keyword(keyword, lastnames, max_results=100):
    print(f"  [OpenAlex] keyword: '{keyword}'")
    results, cursor, fetched = [], "*", 0
    while fetched < max_results:
        per_page = min(50, max_results - fetched)
        data = safe_get("https://api.openalex.org/works", {
            "search": keyword,
            "filter": f"publication_date:>{START_DATE.strftime('%Y-%m-%d')}",
            "per-page": per_page, "cursor": cursor,
            "mailto": CONTACT_EMAIL,
            "select": "id,title,doi,publication_year,publication_date,primary_location,authorships,abstract_inverted_index",
        })
        if not data:
            break
        items = data.get("results", [])
        if not items:
            break
        for w in items:
            row = parse_openalex(w, lastnames)
            if row:
                results.append(row)
        fetched += len(items)
        cursor = data.get("meta", {}).get("next_cursor")
        if not cursor:
            break
        time.sleep(0.3)
    print(f"    → {len(results)} results")
    return results


def openalex_author(scientist, lastnames):
    name = scientist["name"]
    print(f"  [OpenAlex] author: '{name}'")
    data = safe_get("https://api.openalex.org/authors",
                    {"search": name, "per-page": 1, "mailto": CONTACT_EMAIL})
    if not data or not data.get("results"):
        return []
    author_id = data["results"][0]["id"]
    data = safe_get("https://api.openalex.org/works", {
        "filter": f"author.id:{author_id},publication_date:>{START_DATE.strftime('%Y-%m-%d')}",
        "per-page": 50, "mailto": CONTACT_EMAIL,
        "select": "id,title,doi,publication_year,publication_date,primary_location,authorships,abstract_inverted_index",
    })
    if not data:
        return []
    results = [r for w in data.get("results", []) if (r := parse_openalex(w, lastnames))]
    print(f"    → {len(results)} results")
    time.sleep(0.3)
    return results


# ─── 2. CROSSREF ───────────────────────────────────────────────────────────────

def crossref_funder(lastnames):
    print(f"  [CrossRef] NSF award #{NSF_AWARD}")
    data = safe_get("https://api.crossref.org/works", {
        "filter": f"funder:10.13039/100000001,from-pub-date:{START_DATE.strftime('%Y-%m-%d')}",
        "query": f"LEAP {NSF_AWARD}",
        "rows": 100, "mailto": CONTACT_EMAIL,
        "select": "DOI,title,author,published,container-title,abstract,funder",
    })
    if not data:
        return []
    results = []
    for item in data.get("message", {}).get("items", []):
        dp    = (item.get("published") or {}).get("date-parts", [[None]])[0]
        year  = dp[0] if dp else None
        month = dp[1] if len(dp) > 1 else None
        if not year or not in_date_range(str(year), str(month) if month else None):
            continue
        doi   = item.get("DOI") or ""
        title = (item.get("title") or [""])[0]
        journal = (item.get("container-title") or [""])[0]
        link  = f"https://doi.org/{doi}" if doi else ""
        raw   = item.get("author") or []
        authors = [f"{a.get('given','')} {a.get('family','')}".strip() for a in raw]
        first_last = extract_last_name(authors[0]) if authors else ""
        abstract = item.get("abstract") or ""

        # Award-level match
        funders    = item.get("funder") or []
        award_hit  = any(NSF_AWARD in str(aw)
                         for f in funders for aw in (f.get("award") or []))
        leap_ack   = "Yes" if award_hit else detect_leap_ack(f"{title} {abstract}")
        if leap_ack == "No":
            for a in authors:
                if extract_last_name(a).lower() in lastnames:
                    leap_ack = "Likely"
                    break

        notes = f"CrossRef/NSF funder. Authors: {'; '.join(authors[:5])}"
        results.append(make_row(first_last, title, link, journal,
                                notes, leap_ack, year, month, authors))
    print(f"    → {len(results)} results")
    return results


# ─── 3. NASA ADS ───────────────────────────────────────────────────────────────

def ads_query(query_str, lastnames, max_results=100):
    if not ADS_API_TOKEN:
        return []
    print(f"  [NASA ADS] '{query_str}'")
    data = safe_get(
        "https://api.adsabs.harvard.edu/v1/search/query",
        params={
            "q": f"({query_str}) pubdate:[{START_DATE.strftime('%Y-%m-%d')} TO 9999-12-31]",
            "fl": "title,author,year,pubdate,doi,identifier,pub,abstract",
            "rows": max_results, "sort": "date desc",
        },
        headers={"Authorization": f"Bearer {ADS_API_TOKEN}"}
    )
    if not data:
        return []
    results = []
    for doc in data.get("response", {}).get("docs", []):
        title   = (doc.get("title") or [""])[0]
        authors = doc.get("author") or []
        year    = doc.get("year")
        pub_date = doc.get("pubdate") or ""
        month   = int(pub_date[5:7]) if len(pub_date) >= 7 else None
        if not year or not in_date_range(str(year), str(month) if month else None):
            continue
        dois    = doc.get("doi") or []
        doi     = dois[0] if dois else ""
        link    = f"https://doi.org/{doi}" if doi else \
                  f"https://ui.adsabs.harvard.edu/abs/{(doc.get('identifier') or [''])[0]}"
        journal  = doc.get("pub") or ""
        abstract = doc.get("abstract") or ""
        first_last = extract_last_name(authors[0]) if authors else ""

        leap_ack = detect_leap_ack(f"{title} {abstract}")
        if leap_ack == "No":
            for a in authors:
                if extract_last_name(a).lower() in lastnames:
                    leap_ack = "Likely"
                    break

        notes = f"NASA ADS. Authors: {'; '.join(authors[:5])}"
        results.append(make_row(first_last, title, link, journal,
                                notes, leap_ack, year, month, authors))
    print(f"    → {len(results)} results")
    time.sleep(0.5)
    return results


def ads_author(scientist, lastnames):
    if not ADS_API_TOKEN:
        return []
    parts = scientist["name"].strip().split()
    last  = parts[-1]
    init  = parts[0][0] if parts else ""
    return ads_query(f'author:"{last}, {init}"', lastnames, max_results=50)


# ─── 4. ORCID ──────────────────────────────────────────────────────────────────

def orcid_works(scientist, lastnames):
    orcid = scientist.get("orcid")
    if not orcid:
        return []
    name = scientist["name"]
    print(f"  [ORCID] {name}  ({orcid})")
    data = safe_get(
        f"https://pub.orcid.org/v3.0/{orcid}/works",
        headers={"Accept": "application/json"}
    )
    if not data:
        return []
    results = []
    for group in (data.get("group") or []):
        summaries = group.get("work-summary") or []
        if not summaries:
            continue
        s = summaries[0]
        pub_year  = (s.get("publication-date") or {}).get("year",  {})
        pub_month = (s.get("publication-date") or {}).get("month", {})
        year  = int(pub_year.get("value"))  if pub_year  and pub_year.get("value")  else None
        month = int(pub_month.get("value")) if pub_month and pub_month.get("value") else None
        if not year or not in_date_range(str(year), str(month) if month else None):
            continue
        title   = ((s.get("title") or {}).get("title") or {}).get("value") or ""
        journal = ((s.get("journal-title") or {}).get("value")) or ""
        doi     = next(
            (e.get("external-id-value") for e in
             (s.get("external-ids") or {}).get("external-id", [])
             if e.get("external-id-type") == "doi"), ""
        )
        link     = f"https://doi.org/{doi}" if doi else ""
        leap_ack = detect_leap_ack(title) or "Likely"   # ORCID record = known LEAP scientist
        notes    = f"ORCID record for {name}"
        results.append(make_row(extract_last_name(name), title, link, journal,
                                notes, leap_ack, year, month, [name]))
    print(f"    → {len(results)} results")
    time.sleep(0.3)
    return results


# ─── 5. arXiv ──────────────────────────────────────────────────────────────────

def arxiv_search(query, lastnames, max_results=50):
    print(f"  [arXiv] '{query}'")
    try:
        r = requests.get("http://export.arxiv.org/api/query", params={
            "search_query": query, "start": 0,
            "max_results": max_results,
            "sortBy": "submittedDate", "sortOrder": "descending",
        }, headers={"User-Agent": f"LEAP-Scraper/2.0 (mailto:{CONTACT_EMAIL})"}, timeout=20)
        r.raise_for_status()
        root = ET.fromstring(r.text)
    except Exception as e:
        print(f"    arXiv error: {e}")
        return []

    ns = {"atom": "http://www.w3.org/2005/Atom"}
    results = []
    for entry in root.findall("atom:entry", ns):
        title     = (entry.findtext("atom:title", namespaces=ns) or "").replace("\n"," ").strip()
        published = entry.findtext("atom:published", namespaces=ns) or ""
        year  = int(published[:4])  if len(published) >= 4 else None
        month = int(published[5:7]) if len(published) >= 7 else None
        if not year or not in_date_range(str(year), str(month) if month else None):
            continue
        link = entry.findtext("atom:id", namespaces=ns) or ""
        for alt in entry.findall("atom:link", ns):
            if alt.get("title") == "doi":
                link = alt.get("href") or link
        authors  = [a.findtext("atom:name", namespaces=ns) or "" for a in entry.findall("atom:author", ns)]
        abstract = (entry.findtext("atom:summary", namespaces=ns) or "").strip()
        first_last = extract_last_name(authors[0]) if authors else ""

        leap_ack = detect_leap_ack(f"{title} {abstract}")
        if leap_ack == "No":
            for a in authors:
                if extract_last_name(a).lower() in lastnames:
                    leap_ack = "Likely"
                    break

        notes = f"arXiv preprint. Authors: {'; '.join(authors[:5])}"
        results.append(make_row(first_last, title, "arXiv preprint",
                                notes, leap_ack, year, month, authors))
    print(f"    → {len(results)} results")
    time.sleep(1)
    return results


def arxiv_author(scientist, lastnames):
    parts = scientist["name"].strip().split()
    last  = parts[-1]
    first = parts[0] if parts else ""
    q = (f'au:"{last}_{first[0]}" AND '
         f'(abs:climate OR abs:ocean OR abs:atmosphere OR abs:"machine learning")')
    return arxiv_search(q, lastnames, max_results=20)


# ─── 6. SEMANTIC SCHOLAR ───────────────────────────────────────────────────────

def semantic_scholar(keyword, lastnames, max_results=50):
    print(f"  [SemanticScholar] '{keyword}'")
    data = safe_get("https://api.semanticscholar.org/graph/v1/paper/search", {
        "query": keyword, "limit": min(max_results, 100),
        "fields": "title,authors,year,venue,externalIds,abstract,publicationDate",
    })
    if not data:
        return []
    results = []
    for item in data.get("data", []):
        pub_date = item.get("publicationDate") or ""
        year  = int(pub_date[:4])  if len(pub_date) >= 4 else item.get("year")
        month = int(pub_date[5:7]) if len(pub_date) >= 7 else None
        if not year or not in_date_range(str(year), str(month) if month else None):
            continue
        title   = item.get("title") or ""
        doi     = (item.get("externalIds") or {}).get("DOI") or ""
        link    = f"https://doi.org/{doi}" if doi else ""
        journal = item.get("venue") or ""
        authors = [a.get("name","") for a in (item.get("authors") or [])]
        first_last = extract_last_name(authors[0]) if authors else ""
        abstract = item.get("abstract") or ""

        leap_ack = detect_leap_ack(f"{title} {abstract}")
        if leap_ack == "No":
            for a in authors:
                if extract_last_name(a).lower() in lastnames:
                    leap_ack = "Likely"
                    break

        notes = f"SemanticScholar. Authors: {'; '.join(authors[:5])}"
        results.append(make_row(first_last, title, link, journal,
                                notes, leap_ack, year, month, authors))
    print(f"    → {len(results)} results")
    time.sleep(1)
    return results


# ─── FULL-TEXT ACKNOWLEDGMENT VERIFICATION (PDF-FIRST) ────────────────────────
#
# For every paper marked "Likely" or "No", download the full PDF and search
# the entire document for LEAP signals. No HTML scraping — PDFs are consistent
# across all publishers and don't break when websites redesign.
#
# Pipeline for each paper (tried in order until one succeeds):
#   1. Unpaywall API      → finds OA PDF URL for any DOI (free, no key)
#   2. Publisher PDF URLs → direct PDF construction for known OA publishers:
#                           arXiv, EarthArXiv, ESSOAr, Copernicus, PLOS
#   3. DOI landing page   → follow redirects, check if final URL is a PDF
#
# PDF text extraction:
#   - Uses pdfplumber to extract all text from every page
#   - Searches the ENTIRE document (not just a guessed section) so nothing
#     is missed due to non-standard headings like "Financial Support",
#     "Funding Information", "Support", etc.
#   - Also logs which page the signal was found on for traceability
#
# LEAP signals searched (case-insensitive):
#   "learning the earth with ai and physics"
#   "leap science technology center" / "leap center" / "leap stc"
#   "2019625" / "ags-2019625" / "ags 2019625"
#   "nsf stc" / "science and technology center"
# ──────────────────────────────────────────────────────────────────────────────

# All signals to search for — case-insensitive, searched across full PDF text
LEAP_ACK_SIGNALS = [
    "learning the earth with ai and physics",
    "leap science technology center",
    "leap center",
    "leap stc",
    "nsf stc",
    "science and technology center",
    NSF_AWARD,            # "2019625"
    f"nsf {NSF_AWARD}",   # "nsf 2019625"
    "ags-2019625",
    "ags 2019625",
]

# Seconds to wait between PDF fetch requests (be polite to servers)
FULLTEXT_DELAY = 2.0

# Max papers to attempt PDF verification for. Set to None to process all.
# Recommended: start with 50 to test, then set None for full run.
FULLTEXT_MAX = None

# Request headers — identifies the scraper politely
_PDF_HEADERS = {
    "User-Agent": (
        f"LEAP-PublicationScraper/4.0 (NSF Award {NSF_AWARD}; "
        f"mailto:{CONTACT_EMAIL}; research use only)"
    ),
    "Accept": "application/pdf,*/*",
}


def _leap_in_text(text):
    """Return (found: bool, signal: str) — which signal matched, or ''."""
    lower = (text or "").lower()
    for sig in LEAP_ACK_SIGNALS:
        if sig in lower:
            return True, sig
    return False, ""


def _extract_pdf_text(content_bytes):
    """
    Extract all text from a PDF byte string using pdfplumber.
    Returns (full_text: str, page_count: int).
    Tries pymupdf as fallback if pdfplumber fails.
    """
    if not content_bytes:
        return "", 0

    # ── pdfplumber (primary) ──────────────────────────────────────────────────
    if PDF_SUPPORT:
        try:
            with pdfplumber.open(io.BytesIO(content_bytes)) as pdf:
                pages = []
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        pages.append(t)
                return "\n".join(pages), len(pdf.pages)
        except Exception:
            pass

    # ── pymupdf fallback ─────────────────────────────────────────────────────
    try:
        import fitz  # pymupdf
        doc = fitz.open(stream=content_bytes, filetype="pdf")
        pages = [doc[i].get_text() for i in range(len(doc))]
        return "\n".join(pages), len(doc)
    except Exception:
        pass

    return "", 0


def _download_pdf(url, timeout=30):
    """
    Download a URL and return bytes if it is a PDF, else None.
    Follows redirects. Checks Content-Type header.
    """
    try:
        r = requests.get(url, headers=_PDF_HEADERS, timeout=timeout,
                         allow_redirects=True)
        r.raise_for_status()
        ct = r.headers.get("Content-Type", "").lower()
        # Accept if Content-Type says PDF, or URL ends in .pdf, or starts with PDF magic bytes
        is_pdf = (
            "pdf" in ct
            or url.lower().endswith(".pdf")
            or r.content[:4] == b"%PDF"
        )
        if is_pdf:
            return r.content
    except Exception:
        pass
    return None


# ── Unpaywall: get all OA PDF URLs for a DOI ─────────────────────────────────

def _unpaywall_pdf_urls(doi):
    """
    Query Unpaywall and return a list of all OA PDF URLs for the DOI,
    best first. Returns [] if none found or on error.
    """
    if not doi:
        return []
    clean = re.sub(r"^https?://doi\.org/", "", doi).strip()
    try:
        r = requests.get(
            f"https://api.unpaywall.org/v2/{clean}",
            params={"email": CONTACT_EMAIL},
            timeout=12,
        )
        r.raise_for_status()
        data = r.json()
    except Exception:
        return []

    pdf_urls = []
    # Best location first
    best = data.get("best_oa_location") or {}
    if best.get("url_for_pdf"):
        pdf_urls.append(best["url_for_pdf"])
    # All other locations
    for loc in (data.get("oa_locations") or []):
        u = loc.get("url_for_pdf")
        if u and u not in pdf_urls:
            pdf_urls.append(u)
    return pdf_urls


# ── Publisher-specific direct PDF URL constructors ───────────────────────────

def _publisher_pdf_urls(link, doi):
    """
    For known fully-open-access publishers, construct direct PDF URLs
    without needing Unpaywall. Returns list of URLs to try.
    """
    urls = []
    link  = link  or ""
    doi   = doi   or ""
    clean_doi = re.sub(r"^https?://doi\.org/", "", doi).strip()

    # arXiv — always OA, PDF at predictable URL
    arxiv_id = re.search(r"(\d{4}\.\d{4,5})(v\d+)?", link)
    if arxiv_id or "arxiv.org" in link:
        aid = (arxiv_id.group(1) if arxiv_id else "")
        if aid:
            urls.append(f"https://arxiv.org/pdf/{aid}")
            urls.append(f"https://arxiv.org/pdf/{aid}.pdf")

    # EarthArXiv — OA preprint, PDF link on page
    if "eartharxiv.org" in link:
        # EarthArXiv PDF URLs follow /repository/object/XXXX/download/
        repo_id = re.search(r"/view/(\d+)", link)
        if repo_id:
            urls.append(f"https://eartharxiv.org/repository/object/{repo_id.group(1)}/download/")

    # ESSOAr — OA preprint
    if "essoar.org" in link:
        essoar_id = re.search(r"/(10\.\d+/\S+)", link)
        if essoar_id:
            urls.append(f"https://www.essoar.org/pdfjs/{essoar_id.group(1)}")

    # Copernicus journals (acp, gmd, esd, etc.) — fully OA, PDF always available
    if "copernicus.org" in link and clean_doi:
        # Copernicus PDF URL pattern: replace last part of DOI with -pdf suffix
        urls.append(f"https://doi.org/{clean_doi}")  # resolves to PDF-capable page

    # PLOS — fully OA
    if "plos.org" in link and clean_doi:
        article_id = clean_doi.replace("/", "%2F")
        urls.append(
            f"https://journals.plos.org/plosone/article/file?id={article_id}&type=printable"
        )

    return urls


# ── Core verification function ────────────────────────────────────────────────

def verify_acknowledgment(row):
    """
    Download the PDF for a paper and search the full text for LEAP signals.
    Returns updated row dict:
      - 'LEAP Funding Acknowledged?' → 'Yes' if confirmed, else unchanged
      - 'Notes' → appended with verification source and matched signal
    """
    link = row.get("Link") or ""
    if not link:
        return row

    # Normalise DOI
    doi = ""
    if "doi.org/" in link:
        doi = link
    elif re.match(r"^10\.", link):
        doi = f"https://doi.org/{link}"

    found      = False
    signal_hit = ""
    source     = ""

    # ── Collect all PDF URLs to try, in priority order ────────────────────────
    pdf_urls_to_try = []

    # 1. Unpaywall (best OA PDF)
    unpaywall_urls = _unpaywall_pdf_urls(doi)
    pdf_urls_to_try.extend(unpaywall_urls)
    time.sleep(0.4)   # polite pause after Unpaywall call

    # 2. Publisher direct URLs (for known OA publishers)
    direct_urls = _publisher_pdf_urls(link, doi)
    for u in direct_urls:
        if u not in pdf_urls_to_try:
            pdf_urls_to_try.append(u)

    # 3. Try the original DOI/link itself (sometimes resolves to a PDF)
    if link not in pdf_urls_to_try:
        pdf_urls_to_try.append(link)

    # ── Try each PDF URL ──────────────────────────────────────────────────────
    for pdf_url in pdf_urls_to_try:
        content = _download_pdf(pdf_url)
        if not content:
            continue

        full_text, n_pages = _extract_pdf_text(content)
        if not full_text:
            continue

        found, signal_hit = _leap_in_text(full_text)
        if found:
            source = pdf_url
            break

        time.sleep(0.3)

    # ── Update row if confirmed ───────────────────────────────────────────────
    if found:
        row = dict(row)
        row["LEAP Funding Acknowledged?"] = "Yes"
        existing = row.get("Notes") or ""
        row["Notes"] = (
            f"{existing} | PDF verified: '{signal_hit}' found in {source[:70]}"
        ).strip(" |")

    return row


# ── Batch runner ─────────────────────────────────────────────────────────────

def run_fulltext_verification(rows):
    """
    Run PDF verification for all rows marked 'Likely' or 'No'.
    Updates rows in-place and returns the modified list.
    Prints live progress so you can see what's happening.
    """
    to_verify = [
        (i, r) for i, r in enumerate(rows)
        if r.get("LEAP Funding Acknowledged?") in ("Likely", "No")
        and r.get("Link")
    ]

    total    = len(to_verify)
    upgraded = 0
    failed   = 0   # no PDF found at all
    not_found = 0  # PDF found but no LEAP signal

    print(f"\n[PDF Verification] {total} papers to check...")
    if FULLTEXT_MAX and total > FULLTEXT_MAX:
        to_verify = to_verify[:FULLTEXT_MAX]
        print(f"  Capped at {FULLTEXT_MAX} — set FULLTEXT_MAX = None to process all")

    for n, (orig_idx, row) in enumerate(to_verify, 1):
        title_short = (row.get("Title") or "No title")[:60]
        print(f"  [{n}/{len(to_verify)}] {title_short}")

        updated = verify_acknowledgment(row)
        time.sleep(FULLTEXT_DELAY)

        if updated["LEAP Funding Acknowledged?"] == "Yes":
            rows[orig_idx] = updated
            upgraded += 1
            sig = ""
            if "PDF verified:" in (updated.get("Notes") or ""):
                sig = updated["Notes"].split("PDF verified:")[-1].strip()[:60]
            print(f"    ✓  LEAP confirmed — {sig}")
        else:
            link = row.get("Link") or ""
            # Distinguish "couldn't get PDF" from "got PDF, LEAP not mentioned"
            not_found += 1

    print(f"\n  ── PDF Verification Summary ──────────────────────────")
    print(f"  Papers checked:          {len(to_verify)}")
    print(f"  Upgraded to Yes:         {upgraded}")
    print(f"  LEAP not found in PDF:   {not_found}")
    print(f"  (remaining Likely/No still need manual review)")

    return rows


# ─── DEDUPLICATION ─────────────────────────────────────────────────────────────

def deduplicate(rows):
    seen     = {}
    priority = {"Yes": 0, "Likely": 1, "No": 2}
    for row in rows:
        key = (row["_work_id"] or row["Title"].lower().strip()[:80]).strip().lower()
        if key not in seen:
            seen[key] = row
        elif priority.get(row["LEAP Funding Acknowledged?"], 2) < \
             priority.get(seen[key]["LEAP Funding Acknowledged?"], 2):
            seen[key] = row
    return list(seen.values())


# ─── EXCEL OUTPUT ──────────────────────────────────────────────────────────────

EXPORT_COLS = [
    "Last Name", "Title", "Link", "Published In",
    "Notes", "LEAP Funding Acknowledged?", "Date", "Citation (NSF Format)"
]
HEADER_COLOR  = "003865"
YES_COLOR     = "C6EFCE"
LIKELY_COLOR  = "FFEB9C"
NO_COLOR      = "FFC7CE"


def write_excel(rows, filepath):
    export = [{c: r.get(c,"") for c in EXPORT_COLS} for r in rows]
    df = pd.DataFrame(export, columns=EXPORT_COLS)
    df.to_excel(filepath, index=False, sheet_name="LEAP Publications")

    wb = load_workbook(filepath)
    ws = wb.active
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR, end_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 32

    for col, w in {"A":16,"B":52,"C":42,"D":28,"E":40,"F":26,"G":12,"H":72}.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    yes_f    = PatternFill("solid", start_color=YES_COLOR,    end_color=YES_COLOR)
    likely_f = PatternFill("solid", start_color=LIKELY_COLOR, end_color=LIKELY_COLOR)
    no_f     = PatternFill("solid", start_color=NO_COLOR,     end_color=NO_COLOR)

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        leap_val = ws.cell(row=idx, column=6).value or ""
        for cell in row:
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = border
            if cell.column == 6:
                if   leap_val == "Yes":    cell.fill = yes_f;    cell.font = Font(name="Arial", size=10, bold=True, color="276221")
                elif leap_val == "Likely": cell.fill = likely_f; cell.font = Font(name="Arial", size=10, bold=True, color="7D5A00")
                else:                      cell.fill = no_f
        ws.row_dimensions[idx].height = 55

    for idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=idx, column=3)
        if (cell.value or "").startswith("http"):
            cell.hyperlink = cell.value
            cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total = len(df)
    for r in [
        ["LEAP Publication Scrape Summary", ""],
        ["Generated",   datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Start Date",  START_DATE.strftime("%B %d, %Y")],
        ["NSF Award",   f"#{NSF_AWARD}"],
        ["", ""],
        ["Total Publications",                         total],
        ["LEAP Explicitly Acknowledged (Yes)",         (df["LEAP Funding Acknowledged?"]=="Yes").sum()],
        ["Likely LEAP – author match (Likely)",        (df["LEAP Funding Acknowledged?"]=="Likely").sum()],
        ["No LEAP Signal – manual review needed (No)", (df["LEAP Funding Acknowledged?"]=="No").sum()],
    ]:
        ws2.append(r)
    ws2.column_dimensions["A"].width = 48
    ws2.column_dimensions["B"].width = 20
    ws2["A1"].font = Font(bold=True, size=14, name="Arial", color=HEADER_COLOR)

    wb.save(filepath)
    print(f"\n  Saved {total} publications to {filepath}")
    return total


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  LEAP Publication Scraper  v2.0")
    print(f"  NSF Award #{NSF_AWARD}  |  From: {START_DATE.strftime('%B %d, %Y')}")
    print("=" * 62)

    scientists = load_scientists(SCIENTISTS_FILE)
    lastnames  = {s["name"].strip().split()[-1].lower() for s in scientists}
    all_rows   = []

    # 1 ── OpenAlex keyword
    print(f"\n[1/6] OpenAlex keyword search ({len(LEAP_KEYWORDS)} queries)...")
    for kw in LEAP_KEYWORDS:
        all_rows.extend(openalex_keyword(kw, lastnames))
        time.sleep(0.5)

    # 2 ── OpenAlex author
    print(f"\n[2/6] OpenAlex author search ({len(scientists)} scientists)...")
    for sci in scientists:
        all_rows.extend(openalex_author(sci, lastnames))

    # 3 ── CrossRef / NSF funder
    print(f"\n[3/6] CrossRef – NSF funder award #{NSF_AWARD}...")
    all_rows.extend(crossref_funder(lastnames))

    # 4 ── NASA ADS
    ads_enabled = bool(ADS_API_TOKEN)
    if ads_enabled:
        print(f"\n[4/6] NASA ADS keyword search ({len(ADS_QUERIES)} queries)...")
        for q in ADS_QUERIES:
            all_rows.extend(ads_query(q, lastnames))
            time.sleep(0.5)
        print(f"      NASA ADS author search ({len(scientists)} scientists)...")
        for sci in scientists:
            all_rows.extend(ads_author(sci, lastnames))
    else:
        print(f"\n[4/6] NASA ADS – SKIPPED (no ADS_API_TOKEN set)")
        print("      Get a free token at: https://ui.adsabs.harvard.edu/user/settings/token")
        print("      Then run:  export ADS_API_TOKEN=your_token")

    # 5 ── ORCID
    orcid_scientists = [s for s in scientists if s.get("orcid")]
    print(f"\n[5/6] ORCID – {len(orcid_scientists)} scientists with ORCID IDs...")
    for sci in orcid_scientists:
        all_rows.extend(orcid_works(sci, lastnames))

    # 6 ── arXiv
    print(f"\n[6/6] arXiv keyword + author search...")
    for q in ARXIV_QUERIES:
        all_rows.extend(arxiv_search(q, lastnames))
    for sci in scientists:
        all_rows.extend(arxiv_author(sci, lastnames))

    # Bonus ── Semantic Scholar
    print(f"\n[Bonus] Semantic Scholar supplementary search...")
    for kw in LEAP_KEYWORDS[:2]:
        all_rows.extend(semantic_scholar(kw, lastnames))

    # Dedup + sort
    print(f"\nRaw results:         {len(all_rows)}")
    deduped = deduplicate(all_rows)
    print(f"After deduplication: {len(deduped)}")

    # ── Full-text acknowledgment verification ─────────────────────────────────
    # Fetches full text for every "Likely" and "No" paper and upgrades to "Yes"
    # if LEAP / award #2019625 is found in the acknowledgments section.
    deduped = run_fulltext_verification(deduped)

    priority = {"Yes": 0, "Likely": 1, "No": 2}
    deduped.sort(key=lambda r: (
        priority.get(r["LEAP Funding Acknowledged?"], 2),
        -(int(r["Date"][:4]) if r["Date"][:4].isdigit() else 0),
        r["Last Name"].lower()
    ))

    write_excel(deduped, OUTPUT_FILE)
    subprocess.run(["gsutil", "cp", OUTPUT_FILE,
                    "gs://leap-persistent/publications/LEAP_Publications.xlsx"], check=True)

    print("\nColumn legend:")
    print("  Green  (Yes)    – LEAP / NSF #2019625 confirmed in full text OR metadata")
    print("  Yellow (Likely) – known LEAP scientist; full text unavailable/paywalled")
    print("  Red    (No)     – no LEAP signal anywhere; manual review recommended")
    print(f"\nTo add/remove scientists: edit {SCIENTISTS_FILE}")
    print("To enable NASA ADS:       export ADS_API_TOKEN=<your_token>")
    print("To limit full-text checks: set FULLTEXT_MAX = 100 (currently:", FULLTEXT_MAX, ")")


if __name__ == "__main__":
    try:
        with ResourceMonitor():     # logs CPU/RAM/network every 30s
            with ProgressLogger():  # logs dask task progress every 30s
                main()
        notify_slack("my_project finished: gs://leap-scratch/me/output.zarr")
    except Exception:
        import traceback
        traceback.print_exc()
        notify_slack("my_project failed — check the logs")
        raise